# how rows with data?
# how many col with data?
# fetch the data

import openpyxl

def GetRow(xlpath, sheet):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.max_row

def GetCol(xlpath, sheet):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.max_column

def ReadData(xlpath, sheet, r, c):
    wb = openpyxl.load_workbook(xlpath)
    sh = wb[sheet]
    return sh.cell(r,c).value